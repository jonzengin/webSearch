import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QTextEdit, QPushButton, QVBoxLayout, QWidget, QMessageBox, QFileDialog, QLabel
import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def adjust_excel_format(file_path):
    """Ensure Excel cells are single-line without wrap text."""
    workbook = load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=False)

    workbook.save(file_path)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Company Data Manager")
        self.setGeometry(100, 100, 600, 600)

        # Selected Excel file
        self.selected_excel_file = None

        # Layout and widgets
        self.layout = QVBoxLayout()
        self.file_label = QLabel("Selected File: No file selected.")
        self.select_file_button = QPushButton("Select Excel File")
        self.text_edit = QTextEdit()
        self.add_button = QPushButton("Process Data and Add to Excel")

        # Arrange widgets
        self.layout.addWidget(self.file_label)
        self.layout.addWidget(self.select_file_button)
        self.layout.addWidget(self.text_edit)
        self.layout.addWidget(self.add_button)

        # Connect signals
        self.select_file_button.clicked.connect(self.select_excel_file)
        self.add_button.clicked.connect(self.process_input)

        # Main widget and layout
        container = QWidget()
        container.setLayout(self.layout)
        self.setCentralWidget(container)

    def select_excel_file(self):
        """Allow user to select an Excel file."""
        file_name, _ = QFileDialog.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xlsx)")
        if file_name:
            self.selected_excel_file = file_name
            self.ensure_columns_exist(file_name)
            self.file_label.setText(f"Selected File: {file_name}")
            QMessageBox.information(self, "File Selected", "Excel file successfully selected!")
        else:
            QMessageBox.warning(self, "Warning", "Please select a valid Excel file!")

    def ensure_columns_exist(self, file_name):
        """Ensure required columns exist in the Excel file."""
        required_columns = [
            "Company Name", "Company Phone", "Company Email", 
            "Industry", "Company Type", "Import/Export Activities", 
            "Company Country", "Company Website", "Operations"
        ]
        if os.path.exists(file_name):
            try:
                df = pd.read_excel(file_name, engine='openpyxl')
                for column in required_columns:
                    if column not in df.columns:
                        df[column] = None
                df.to_excel(file_name, index=False, engine='openpyxl')
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error while checking Excel file: {e}")
        else:
            df = pd.DataFrame(columns=required_columns)
            df.to_excel(file_name, index=False, engine='openpyxl')

    def process_input(self):
        """Process user input and add data to Excel."""
        if not self.selected_excel_file:
            QMessageBox.critical(self, "Error", "Please select an Excel file first!")
            return

        input_text = self.text_edit.toPlainText().strip()
        if not input_text:
            QMessageBox.critical(self, "Error", "Please paste data into the text box!")
            return

        company_data = self.parse_text(input_text)
        if not company_data:
            return  # Error already shown in parse_text

        if self.add_to_excel(company_data):
            QMessageBox.information(self, "Success", "Data successfully added to Excel!")
        else:
            QMessageBox.warning(self, "Warning", "This company already exists in the file!")

    def parse_text(self, input_text):
        """Parse input text and extract company details."""
        try:
            company_data = {
                "Company Name": re.search(r"Company Name: (.+)", input_text).group(1).strip(),
                "Company Phone": re.search(r"Company Phone: (.+)", input_text).group(1).strip(),
                "Company Email": self.extract_optional_field(r"Company Email: (.+)", input_text),
                "Industry": re.search(r"Industry: (.+)", input_text).group(1).strip(),
                "Company Type": re.search(r"Company Type: (.+)", input_text).group(1).strip(),
                "Import/Export Activities": self.extract_optional_field(r"Import/Export Activities: (.+)", input_text),
                "Company Country": re.search(r"Company Country: (.+)", input_text).group(1).strip(),
                "Company Website": self.extract_optional_field(r"Company Website: (.+)", input_text),
                "Operations": re.search(r"Company Operations: (.+)", input_text, re.DOTALL).group(1).strip()
            }
            return company_data
        except AttributeError as e:
            QMessageBox.critical(self, "Error", f"Invalid data format: {e}")
            return None

    def extract_optional_field(self, pattern, text):
        """Extract optional field value or return None if not found."""
        match = re.search(pattern, text)
        return match.group(1).strip() if match else None

    def add_to_excel(self, company_data):
        """Add parsed data to the Excel file."""
        try:
            df = pd.read_excel(self.selected_excel_file, engine='openpyxl')

            if not df[df["Company Name"] == company_data["Company Name"]].empty:
                return False

            new_data = pd.DataFrame([company_data])
            df = pd.concat([df, new_data], ignore_index=True)
            df.to_excel(self.selected_excel_file, index=False, engine='openpyxl')

            adjust_excel_format(self.selected_excel_file)
            return True
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error while adding data to Excel: {e}")
            return False

def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
